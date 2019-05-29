import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from lxml import html
import requests
from copy import copy

# domain= domain name... text value of cell
# returns true if available | false if taken
def checkDomain(domain):
    url='https://www.norid.no/no/?query='+domain+'.no'
    page = requests.get(url)
    tree = html.fromstring(page.content)
    response = tree.xpath('//b/text()')
    if len(response)>0:
        if response[0].endswith('t'):
            return False
        else:
            return True

# Create fill colors
redFill = PatternFill(start_color='ffc7ce',
                   end_color='ffc7ce',
                   fill_type='solid')
greenFill = PatternFill(start_color='c6efce',
                   end_color='c6efce',
                   fill_type='solid')
# load source workbook
wb_in= load_workbook(filename = "domain_names.xlsx", data_only=True)
ws_in= wb_in.active
# create target workbook
wb_out=Workbook()
ws_out=wb_out.active

#loop through cells
for row in range(1, ws_in.max_row):
    for col in range(1, ws_in.max_column):
        source=ws_in.cell(row=row, column=col) # input
        target=ws_out.cell(row=row, column=col) # output
        if source.value: #if not empy
            if col==1 or row==1 or source.fill.patternType: #check if keyword or if already checked
                target.value=source.value #copy
                target.fill=copy(source.fill)4646
            else: #check if domain available
                if checkDomain(source.value):
                    target.fill=greenFill
                    target.value=source.value
                else:
                    target.fill=redFill
                    target.value=source.value
# write file
wb_out.save(filename="domains.xlsx")
